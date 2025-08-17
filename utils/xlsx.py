import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

async def create_report_excel(name: str, rows: list[list]) -> str:
    headers = [
        "usertag", "цена", "во_сколько_начали_идти_заявки", "актуал_заявок",
        "цена_заявки", "пдп_осталось", "цена_пдп_за_72ч", "процент_отписок",
        "пришло_за_10_мин", "пришло_за_28_мин", "пришло_за_60_мин",
        "пришло_за_24ч", "пришло_за_48ч", "пришло_за_72ч"
    ]
    header_titles = {
        "usertag": "@usertag",
        "цена": "цена",
        "во_сколько_начали_идти_заявки": "Начало",
        "актуал_заявок": "Актуал. заявки",
        "цена_заявки": "Цена заявки",
        "пдп_осталось": "ПДП",
        "цена_пдп_за_72ч": "Цена ПДП",
        "процент_отписок": "% отписок",
        "пришло_за_10_мин": "+ За 10 мин",
        "пришло_за_28_мин": "+ За 28 мин",
        "пришло_за_60_мин": "+ За 60 мин",
        "пришло_за_24ч": "+ За 24ч",
        "пришло_за_48ч": "+ За 48ч",
        "пришло_за_72ч": "+ За 72 ч"
    }
    column_widths = {
        "usertag": 22,
        "цена": 17,
        "во_сколько_начали_идти_заявки": 25,
        "актуал_заявок": 18,
        "цена_заявки": 16,
        "пдп_осталось": 15,
        "цена_пдп_за_72ч": 18,
        "процент_отписок": 18,
        "пришло_за_10_мин": 25,
        "пришло_за_28_мин": 25,
        "пришло_за_60_мин": 25,
        "пришло_за_24ч": 25,
        "пришло_за_48ч": 25,
        "пришло_за_72ч": 25
    }

    os.makedirs('files', exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    header_font = Font(size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    cell_font = Font(size=12)
    cell_fill = PatternFill("solid", fgColor="DCE6F1")
    center_align = Alignment(horizontal="center", vertical="center")
    border_side = Side(border_style="thin", color="000000")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Записываем заголовки
    for col_num, key in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_titles[key])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = column_widths.get(key, 15)

    # Записываем данные с вычислениями
    for row_idx, row_data in enumerate(rows, start=2):
        usertag = row_data[0]
        price = row_data[1]
        start_time = row_data[2]
        requests = row_data[3]
        joins = row_data[4]
        m10 = row_data[5]
        m28 = row_data[6]
        m60 = row_data[7]
        m24 = row_data[8]
        m48 = row_data[9]
        m72 = row_data[10]
        m10_j = row_data[11]
        m28_j = row_data[12]
        m60_j = row_data[13]
        m24_j = row_data[14]
        m48_j = row_data[15]
        m72_j = row_data[16]

        price_per_request = price / requests if requests else 0
        price_per_join_72 = price / joins if joins else 0
        unsub_percent = round(int(round((requests - joins) / requests, 4))*4, 2) if requests else 0

        values = [
            usertag, price, start_time, requests,
            round(price_per_request, 2), joins, round(price_per_join_72, 2), unsub_percent,
            f'{m10} з / {m10_j} пдп', f'{m28} з / {m28_j} пдп', f'{m60} з / {m60_j} пдп',
            f'{m24} з / {m24_j} пдп', f'{m48} з / {m48_j} пдп', f'{m72} з / {m72_j} пдп'
        ]

        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=value)
            cell.font = cell_font
            cell.fill = cell_fill
            cell.alignment = center_align
            cell.border = border

    filepath = f"files/{name}.xlsx"
    wb.save(filepath)
    return filepath



import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference


import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


def export_conversion_to_excel(
        privets: list,
        links:   list,
        elements: list,
        out_path: str = "files/стата_по_цепи.xlsx"
):
    """
    Формирует Excel‑отчёт вида

        Звено №1
            Приветка‑А   | Заявки | Конверсия %
            Приветка‑Б   | ...
        Звено №2
            ...

    • конверсия считается от previuos_pair (если предыдущие заявки > 0);
    • если заявок нет → «нет заявок»;
    • если у предыдущей пары заявок 0 → «конверсия = 0».

    Parameters
    ----------
    privets  – [{'id','text',…}]
    links    – [{'element1_id','privet_id','requests','previuos_pair', …}]
    elements – [{'id','channel_name', …}]
    """
    priv_map = {p["id"]: p["text"] for p in privets}
    elem_map = {e["id"]: e["channel_name"] for e in elements}
    link_map = {l["id"]: l for l in links}

    summary = {}
    for link in links:
        zveno = elem_map.get(link["element1_id"], "—")

        if zveno not in summary:
            summary[zveno] = {"приветка": set(), "залито": 0, "перелито": 0}

        priv = priv_map.get(link["privet_id"])
        if priv:
            summary[zveno]["приветка"].add(priv)

        summary[zveno]["залито"] += link["requests"]

        prev = link_map.get(link["previuos_pair"])
        if prev:
            summary[zveno]["перелито"] += prev["requests"]

    result = {
        "приветка": {},
        "залито": {},
        "перелито": {},
        "конверсия": {}
    }

    for zveno, data in summary.items():
        priv_count = len(data["приветка"])
        zalito = data["залито"]
        perelito = data["перелито"]
        konv = round(perelito / zalito, 4) if zalito else "#ДЕЛ/0!"

        result["приветка"][zveno] = priv_count
        result["залито"][zveno] = zalito
        result["перелито"][zveno] = perelito
        result["конверсия"][zveno] = konv

    wb = Workbook()
    ws = wb.active
    ws.title = "Конверсии"

    zvenya = list(result["приветка"].keys())
    ws.cell(row=16, column=1, value="ЗВЕНО")
    for i, z in enumerate(zvenya):
        ws.cell(row=16, column=i + 2, value=z)

    ws.cell(row=17, column=1, value="АКТУАЛЬНАЯ")

    labels = ["приветка", "залито", "перелито", "конверсия"]
    for i, label in enumerate(labels):
        ws.cell(row=18 + i, column=1, value=label)
        for j, z in enumerate(zvenya):
            ws.cell(row=18 + i, column=2 + j, value=result[label][z])

    ws.cell(row=22, column=1, value="АРХИВНЫЕ")
    for i, label in enumerate(labels):
        ws.cell(row=23 + i, column=1, value=label)

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=16, max_row=27, min_col=1, max_col=1 + len(zvenya)):
        for cell in row:
            cell.alignment = center
            if cell.row in [16, 17, 22] or cell.column == 1:
                cell.font = bold

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    return os.path.abspath(out_path)



# ==== пример вызова ====
# path = export_zveno_privets(privets, links, elements)
# print("Excel сохранён:", path)
